# 动物检疫合格证明查询模块
# 调用四川省动物卫生监督所公众查询 API

import json
import logging
import requests
from typing import Dict, List, Optional, Any
from dataclasses import dataclass, field

logger = logging.getLogger(__name__)

# ─────────────────────── 证书类型配置 ─────────────────────── #

CERT_TYPES = {
    "1": {"name": "动物A证",     "categoryId": "bca2277f177f41a6a65d836fff78ffca"},
    "2": {"name": "动物A证（外省）", "categoryId": "7f346e9b82414eb2bf1f3126906abe29"},
    "3": {"name": "动物B证",     "categoryId": "44be5d010ab14aaaa99383eb72875145"},
    "4": {"name": "产品A证",     "categoryId": "4b3c752b81d74d94a6ee43bda3e7a85f"},
    "5": {"name": "产品B证",     "categoryId": "896d60db27154903a7ad413534962f48"},
    "6": {"name": "产品A证（外省）", "categoryId": "39d8b64a776d45af88172c051cf89bf9"},
    "7": {"name": "产品证",      "categoryId": "79aa766f2c6546efbed2ae5eced7adcd"},
}

# ─────────────────────── 查询字段映射（每种类型返回的字段不同） ─────────────────────── #

CERT_FIELDS = {
    "1": ["检疫证号", "货主名称", "货主电话", "动物种类", "数量", "单位", "用途",
          "启运地", "启运地点", "承运人", "承运人电话", "运载方式", "运载工具牌号",
          "目的地", "到达地点", "有效期", "检疫证状态", "签发日期"],
    "2": ["检疫证号", "货主名称", "货主电话", "动物种类", "数量", "单位", "用途",
          "启运地", "启运地点", "承运人", "承运人电话", "运载方式", "运载工具牌号",
          "目的地", "到达地点", "有效期", "检疫证状态", "签发日期"],
    "3": ["检疫证号", "货主名称", "货主电话", "动物种类", "数量", "单位", "用途",
          "启运地", "启运地点", "启运点类型", "目的地", "到达地点", "到达点类型",
          "承运人", "承运人电话", "运载方式", "运载工具牌号", "耳标号",
          "检疫员", "检疫申报点", "检疫机构", "检疫证状态", "打印状态", "签发日期"],
    "4": ["检疫证号", "货主名称", "货主电话", "数量", "单位",
          "启运地", "启运地点", "承运人", "承运人电话", "运载方式", "运载工具牌号",
          "目的地", "到达地点", "检疫证状态", "签发日期"],
    "5": ["检疫证号", "货主名称", "货主电话", "数量", "单位",
          "启运地", "启运地点", "目的地", "到达地点", "检疫证状态", "签发日期"],
    "6": ["检疫证号", "货主名称", "货主电话", "数量", "单位",
          "启运地", "启运地点", "承运人", "承运人电话", "运载方式", "运载工具牌号",
          "目的地", "到达地点", "检疫证状态", "签发日期"],
    "7": ["签发日期", "生产单位区划", "产品生产单位名称", "申报点", "检疫证号"],
}

# ─────────────────────── 字段路径映射 ─────────────────────── #

FIELD_PATHS = {
    # 动物A证（跨省调运）— 完整字段
    "1": [
        ("FactoryCode", "检疫证号"),
        ("Owner.Name", "货主名称"),
        ("OwnerTel", "货主电话"),
        ("AnimalType.Name", "动物种类"),
        ("Amount", "数量"),
        ("Unit.Name", "单位"),
        ("UsageType.Name", "用途"),
        ("StartingPlaceRegion.RegionFullName", "启运地"),
        ("StartingPlaceName", "启运地点"),
        ("Carrier.Name", "承运人"),
        ("CarrierTel", "承运人电话"),
        ("TransportType.Name", "运载方式"),
        ("TransportNumber", "运载工具牌号"),
        ("DestinationPlaceRegion.RegionFullName", "目的地"),
        ("DestinationPlaceAddress", "到达地点"),
        ("ValidityDays.Name", "有效期"),
        ("Status.Name", "检疫证状态"),
        ("DateOfIssue", "签发日期"),
    ],
    # 动物A证（外省）— Owner/Carrier 为字符串
    "2": [
        ("FactoryCode", "检疫证号"),
        ("Owner", "货主名称"),
        ("OwnerTel", "货主电话"),
        ("AnimalType.Name", "动物种类"),
        ("Amount", "数量"),
        ("Unit.Name", "单位"),
        ("UsageType.Name", "用途"),
        ("StartingPlaceRegion.RegionFullName", "启运地"),
        ("StartingPlaceName", "启运地点"),
        ("Carrier", "承运人"),
        ("CarrierTel", "承运人电话"),
        ("TransportType.Name", "运载方式"),
        ("TransportNumber", "运载工具牌号"),
        ("DestinationPlaceRegion.RegionFullName", "目的地"),
        ("DestinationPlaceAddress", "到达地点"),
        ("ValidityDays.Name", "有效期"),
        ("Status.Name", "检疫证状态"),
        ("DateOfIssue", "签发日期"),
    ],
    # 动物B证（省内）— 有承运人但字段名小写
    "3": [
        ("FactoryCode", "检疫证号"),
        ("Owner.Name", "货主名称"),
        ("OwnerTel", "货主电话"),
        ("AnimalType.Name", "动物种类"),
        ("Amount", "数量"),
        ("Unit.Name", "单位"),
        ("UsageType.Name", "用途"),
        ("StartingPlaceRegion.RegionFullName", "启运地"),
        ("StartingPlaceName", "启运地点"),
        ("StartingPlaceType.Name", "启运点类型"),
        ("DestinationPlaceRegion.RegionFullName", "目的地"),
        ("DestinationPlaceName", "到达地点"),
        ("DestinationPlaceType.Name", "到达点类型"),
        ("carrier", "承运人"),
        ("carriertel", "承运人电话"),
        ("trafficname", "运载方式"),
        ("TransportNumber", "运载工具牌号"),
        ("EarTags", "耳标号"),
        ("QuarantineOfficer.Name", "检疫员"),
        ("QUADECPoint.Name", "检疫申报点"),
        ("AgencyJson.Name", "检疫机构"),
        ("Status.Name", "检疫证状态"),
        ("PrintStatus.Name", "打印状态"),
        ("DateOfIssue", "签发日期"),
    ],
    # 产品A证（跨省）— 无动物种类
    "4": [
        ("FactoryCode", "检疫证号"),
        ("Owner.Name", "货主名称"),
        ("OwnerTel", "货主电话"),
        ("Amount", "数量"),
        ("Unit.Name", "单位"),
        ("StartingPlaceRegion.RegionFullName", "启运地"),
        ("StartingPlaceName", "启运地点"),
        ("Carrier.Name", "承运人"),
        ("CarrierTel", "承运人电话"),
        ("TransportType.Name", "运载方式"),
        ("TransportNumber", "运载工具牌号"),
        ("DestinationPlaceRegion.RegionFullName", "目的地"),
        ("DestinationPlaceAddress", "到达地点"),
        ("Status.Name", "检疫证状态"),
        ("DateOfIssue", "签发日期"),
    ],
    # 产品B证（省内）— 无承运人
    "5": [
        ("FactoryCode", "检疫证号"),
        ("Owner.Name", "货主名称"),
        ("OwnerTel", "货主电话"),
        ("Amount", "数量"),
        ("Unit.Name", "单位"),
        ("StartingPlaceRegion.RegionFullName", "启运地"),
        ("StartingPlaceName", "启运地点"),
        ("DestinationPlaceRegion.RegionFullName", "目的地"),
        ("DestinationPlaceAddress", "到达地点"),
        ("Status.Name", "检疫证状态"),
        ("DateOfIssue", "签发日期"),
    ],
    # 产品A证（外省）— Carrier 为字符串
    "6": [
        ("FactoryCode", "检疫证号"),
        ("Owner.Name", "货主名称"),
        ("OwnerTel", "货主电话"),
        ("Amount", "数量"),
        ("Unit.Name", "单位"),
        ("StartingPlaceRegion.RegionFullName", "启运地"),
        ("StartingPlaceName", "启运地点"),
        ("Carrier", "承运人"),
        ("CarrierTel", "承运人电话"),
        ("TransportType.Name", "运载方式"),
        ("TransportNumber", "运载工具牌号"),
        ("DestinationPlaceRegion.RegionFullName", "目的地"),
        ("DestinationPlaceAddress", "到达地点"),
        ("Status.Name", "检疫证状态"),
        ("DateOfIssue", "签发日期"),
    ],
    # 产品证（独立结构，snake_case 字段名）
    "7": [
        ("date_of_issue", "签发日期"),
        ("create_company_region.RegionFullName", "生产单位区划"),
        ("create_company_name", "产品生产单位名称"),
        ("issued_organization_name", "申报点"),
        ("factory_code", "检疫证号"),
    ],
}

# API 配置
API_URL = "https://www.scahi.org.cn/v3.0/api/objects/query/first"
PARTITION_ID = "0b1d178c499043a2aeeef591a3d8f03d"
PART_ID = "d5896b31964e425382df52f655dedfc2"


@dataclass
class CertQueryResult:
    """检疫证查询结果"""
    success: bool
    cert_type: str = ""
    factory_code: str = ""
    fields: Dict[str, str] = field(default_factory=dict)
    raw_data: Optional[Dict] = None
    error_msg: str = ""


def _get_nested_value(obj: dict, path: str) -> str:
    """从嵌套对象中按路径取值"""
    if not obj or not path:
        return "-"
    keys = path.split(".")
    result = obj
    for key in keys:
        if isinstance(result, dict) and key in result:
            result = result[key]
        else:
            return "-"
    return str(result) if result else "-"


def query_cert(factory_code: str, cert_type: str = "1") -> CertQueryResult:
    """
    查询检疫合格证明

    Args:
        factory_code: 检疫证号（10-11位印刷号或系统生成号）
        cert_type: 证书类型（"1"~"7"）

    Returns:
        CertQueryResult 查询结果
    """
    # 验证参数
    cert_type = str(cert_type)
    if cert_type not in CERT_TYPES:
        return CertQueryResult(
            success=False,
            error_msg=f"不支持的证书类型: {cert_type}"
        )

    factory_code = factory_code.strip()
    if len(factory_code) < 10:
        return CertQueryResult(
            success=False,
            error_msg="请输入10位或11位编号"
        )

    category_id = CERT_TYPES[cert_type]["categoryId"]
    cert_name = CERT_TYPES[cert_type]["name"]

    # 构造查询参数
    rk = "factory_code" if cert_type == "7" else "FactoryCode"
    info = {"_PartId": PART_ID}
    info[rk] = f"@=@{factory_code}"

    payload = {
        "dataArgs": [info],
        "orderBy": ["CreateAt|1"],
        "limit": 1
    }

    url = f"{API_URL}?categoryId={category_id}&partitionId={PARTITION_ID}"

    try:
        logger.info(f"[CertQuery] 查询 {cert_name}: {factory_code}")
        resp = requests.post(
            url,
            json=payload,
            headers={
                "Content-Type": "application/json",
                "Referer": "https://www.scahi.org.cn/",
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/131.0.0.0 Safari/537.36"
                ),
                "Accept": "application/json, text/plain, */*",
            },
            timeout=15
        )
        resp.raise_for_status()
        data = resp.json()
    except requests.RequestException as e:
        logger.error(f"[CertQuery] 网络请求失败: {e}")
        return CertQueryResult(
            success=False,
            error_msg=f"网络请求失败: {str(e)}"
        )
    except json.JSONDecodeError as e:
        logger.error(f"[CertQuery] 响应解析失败: {e}")
        return CertQueryResult(
            success=False,
            error_msg="服务器返回数据格式异常"
        )

    # 解析结果
    result_obj = data.get("Result")
    if not result_obj:
        return CertQueryResult(
            success=False,
            cert_type=cert_name,
            factory_code=factory_code,
            error_msg="未查询到数据"
        )

    # 提取字段
    fields = {}
    paths = FIELD_PATHS.get(cert_type, [])
    for path, label in paths:
        fields[label] = _get_nested_value(result_obj, path)

    return CertQueryResult(
        success=True,
        cert_type=cert_name,
        factory_code=factory_code,
        fields=fields,
        raw_data=result_obj,
    )


def get_cert_types() -> List[Dict[str, str]]:
    """获取可用的证书类型列表"""
    return [{"key": k, "name": v["name"]} for k, v in CERT_TYPES.items()]


def export_to_excel(results: List[CertQueryResult], file_path: str) -> bool:
    """将查询结果导出为 Excel 文件"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "检疫证查询结果"

        # 表头样式
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="409EFF", end_color="409EFF", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 写入表头
        all_headers = ["序号", "证书类型", "检疫证号"]
        # 收集所有字段名
        field_names = set()
        for r in results:
            if r.success:
                field_names.update(r.fields.keys())
        all_headers.extend(sorted(field_names))

        for col, header in enumerate(all_headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 写入数据
        for row_idx, result in enumerate(results, 2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1).border = thin_border
            ws.cell(row=row_idx, column=2, value=result.cert_type).border = thin_border
            ws.cell(row=row_idx, column=3, value=result.factory_code).border = thin_border

            for col_idx, header in enumerate(sorted(field_names), 4):
                value = result.fields.get(header, "") if result.success else result.error_msg
                ws.cell(row=row_idx, column=col_idx, value=value).border = thin_border

        # 自动调整列宽
        for col in range(1, len(all_headers) + 1):
            max_width = 0
            for row in range(1, len(results) + 2):
                cell_value = str(ws.cell(row=row, column=col).value or "")
                max_width = max(max_width, len(cell_value))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = min(max_width + 4, 50)

        wb.save(file_path)
        logger.info(f"[CertQuery] 导出成功: {file_path}")
        return True
    except Exception as e:
        logger.error(f"[CertQuery] 导出失败: {e}")
        return False
