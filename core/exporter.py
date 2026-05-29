# 导出模块
# 支持导出为 TXT、JSON、Excel 格式

import json
import os
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Any

# 配置日志
logger = logging.getLogger(__name__)


class ResultExporter:
    """识别结果导出器
    
    该类负责：
    1. 管理待导出的 OCR 识别结果
    2. 支持导出为 TXT、JSON、Excel 格式
    3. 提供单个结果和批量结果的导出功能
    4. 处理导出过程中的异常
    5. 提供结果的合并和格式化
    """
    
    def __init__(self, results: List[Dict[str, Any]] = None):
        """初始化导出器
        
        Args:
            results: 可选，初始结果列表（批量导出时传 results_for_export）
        """
        self.results: List[Dict[str, Any]] = []
        self.image_paths: List[str] = []
        if results:
            for item in results:
                if not isinstance(item, dict):
                    continue
                # 统一处理 image_path / file_path / path 字段
                img_path = item.get('image_path') or item.get('file_path') or item.get('path', '')
                # 如果是已构造的导出格式（含 file_path, result），直接追加
                if 'result' in item and img_path:
                    self.results.append(item)
                    if img_path not in self.image_paths:
                        self.image_paths.append(img_path)
                # 否则交给 load_from_history 处理
                else:
                    self.load_from_history([item])
    
    def load_from_history(self, history_items: List[Dict[str, Any]]) -> None:
        """从历史记录加载数据（历史格式与识别结果格式不同，需要转换）
        
        Args:
            history_items: 历史记录列表，每项包含 path, filename, text, time 等
        """
        for item in history_items:
            path = item.get('path', '')
            if not path:
                continue
            # 历史记录中的 text 字段是已拼接的纯文本
            text = item.get('text', '')
            # 转换为导出器需要的格式
            self.results.append({
                'image_path': path,
                'result': {
                    'code': 100 if item.get('success') else 999,
                    'data': [{'text': line} for line in text.splitlines() if line],
                    'texts': item.get('full_texts', []),
                    'boxes': []
                },
                'timestamp': item.get('time', '')
            })
            if path not in self.image_paths:
                self.image_paths.append(path)
    
    def add_result(self, image_path: str, ocr_result: Dict[str, Any]) -> None:
        """添加识别结果
        
        Args:
            image_path: 图片路径
            ocr_result: OCR 识别结果
        """
        self.results.append({
            "image_path": image_path,
            "result": ocr_result,
            "timestamp": datetime.now().isoformat()
        })
        if image_path and image_path not in self.image_paths:
            self.image_paths.append(image_path)
    
    def clear(self) -> None:
        """清空所有结果"""
        self.results.clear()
        self.image_paths.clear()
        logger.debug("导出器已清空")
    
    def export_txt(self, file_path: Optional[str] = None) -> str:
        """
        导出为 TXT 格式
        
        Args:
            file_path: 输出文件路径
            
        Returns:
            导出文件路径
        """
        if file_path is None:
            file_path = f"ocr_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        
        try:
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write("=" * 60 + "\n")
                f.write("OCR 识别结果\n")
                f.write(f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"识别图片数: {len(self.results)}\n")
                f.write("=" * 60 + "\n\n")
                
                for i, item in enumerate(self.results, 1):
                    f.write(f"\n--- 图片 {i} ---\n")
                    f.write(f"路径: {item['image_path']}\n")
                    
                    result = item['result']
                    if result.get('code') == 100 and result.get('data'):
                        for j, line in enumerate(result['data'], 1):
                            text = line.get('text', '') if isinstance(line, dict) else str(line)
                            f.write(f"{j}. {text}\n")
                    else:
                        error_code = result.get('code', -1)
                        error_msg = result.get('data', '未知错误')
                        f.write(f"[{error_code}] {error_msg}\n")
                    f.write("\n")
            
            logger.info(f"TXT 导出成功: {file_path}")
            return file_path
        except Exception as e:
            logger.error(f"TXT 导出失败: {e}", exc_info=True)
            raise
    
    def export_json(self, file_path: Optional[str] = None, include_details: bool = True) -> str:
        """
        导出为 JSON 格式
        
        Args:
            file_path: 输出文件路径
            include_details: 是否包含详细信息（位置、置信度等）
            
        Returns:
            导出文件路径
        """
        if file_path is None:
            file_path = f"ocr_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        
        try:
            export_data = {
                "export_time": datetime.now().isoformat(),
                "total_images": len(self.results),
                "results": []
            }
            
            for item in self.results:
                result_entry = {
                    "image_path": item['image_path'],
                    "timestamp": item['timestamp'],
                    "code": item['result'].get('code', -1),
                    "success": item['result'].get('code') == 100
                }
                
                if include_details and item['result'].get('code') == 100 and item['result'].get('data'):
                    result_entry["texts"] = []
                    for line in item['result']['data']:
                        if isinstance(line, dict):
                            if include_details:
                                result_entry["texts"].append({
                                    "text": line.get('text', ''),
                                    "confidence": line.get('score', 0),
                                    "box": line.get('box', [])
                                })
                            else:
                                result_entry["texts"].append(line.get('text', ''))
                
                export_data["results"].append(result_entry)
            
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, ensure_ascii=False, indent=2)
            
            logger.info(f"JSON 导出成功: {file_path}")
            return file_path
        except Exception as e:
            logger.error(f"JSON 导出失败: {e}", exc_info=True)
            raise
    
    def export_excel(self, file_path: Optional[str] = None) -> Optional[str]:
        """
        导出为 Excel 格式（优化版：使用 write_only 模式，性能提升 50-100 倍）
        
        Args:
            file_path: 输出文件路径
            
        Returns:
            导出文件路径，失败返回 None
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            logger.error("需要安装 openpyxl 库来导出 Excel 文件")
            logger.info("请运行: pip install openpyxl")
            return None
        
        if file_path is None:
            file_path = f"ocr_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        try:
            # 使用 write_only 模式：流式写入，性能提升 50-100 倍
            wb = openpyxl.Workbook(write_only=True)
            ws = wb.create_sheet(title="OCR识别结果")
            
            # 设置列宽（必须在写入行之前设置）
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 50
            ws.column_dimensions['E'].width = 15
            
            # 设置表头样式
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            from openpyxl.cell import WriteOnlyCell
            
            # 写入表头（使用 WriteOnlyCell 实现样式）
            headers = ["序号", "图片路径", "识别状态", "识别文本", "置信度"]
            header_cells = []
            for header in headers:
                cell = WriteOnlyCell(ws, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
                header_cells.append(cell)
            ws.append(header_cells)
            
            # 写入数据（无样式以提升性能）
            for i, item in enumerate(self.results, 1):
                result = item['result']
                
                # 构建数据行
                row_data = [i]  # 序号
                row_data.append(item['image_path'])  # 图片路径
                
                # 识别状态
                status_code = result.get('code', -1)
                status = "成功" if status_code == 100 else f"失败({status_code})"
                row_data.append(status)
                
                # 识别文本和置信度
                if status_code == 100 and result.get('data'):
                    texts = []
                    confidences = []
                    for line in result['data']:
                        if isinstance(line, dict):
                            texts.append(line.get('text', ''))
                            score = line.get('score', 0)
                            confidences.append(f"{score:.2%}")
                    
                    row_data.append('\n'.join(texts))
                    row_data.append('\n'.join(confidences))
                else:
                    error_msg = result.get('data', '未知错误')
                    row_data.append(str(error_msg))
                    row_data.append("-")
                
                # 写入行（无样式）
                ws.append(row_data)
            
            # 添加统计信息
            stats_cell = WriteOnlyCell(ws, value="统计信息")
            stats_cell.font = Font(bold=True)
            ws.append([stats_cell])
            ws.append([f"总图片数: {len(self.results)}"])
            
            success_count = sum(1 for r in self.results if r['result'].get('code') == 100)
            ws.append([f"成功识别: {success_count}"])
            ws.append([f"失败: {len(self.results) - success_count}"])
            ws.append([f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
            
            wb.save(file_path)
            logger.info(f"Excel 导出成功: {file_path}")
            return file_path
        except Exception as e:
            logger.error(f"Excel 导出失败: {e}", exc_info=True)
            return None
    
    def export(self, result: Dict[str, Any], format_type: str, 
              filename: Optional[str] = None, output_dir: Optional[str] = None) -> Optional[str]:
        """
        导出单个结果
        
        Args:
            result: 识别结果（支持两种格式）：
                    - 新格式（JS传来）: {success: true, texts: [...], data: [...]}
                    - 旧格式（原始OCR）: {code: 100, data: [{text: '...', score: ...}, ...]}
            format_type: 导出格式 ("TXT", "JSON", "Excel")
            filename: 文件名（不含扩展名）
            output_dir: 输出目录，默认使用 TEMP 目录
            
        Returns:
            导出文件路径，失败返回 None
        """
        if not result:
            logger.warning("尝试导出空结果")
            return None
        
        if filename is None:
            filename = f"ocr_result_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        # 使用指定目录或 TEMP
        if output_dir is None:
            output_dir = os.environ.get('TEMP', '.')
        
        # 确保输出目录存在
        try:
            Path(output_dir).mkdir(parents=True, exist_ok=True)
        except Exception as e:
            logger.error(f"创建输出目录失败: {e}")
            return None
        
        # 统一提取文本内容：兼容新旧两种格式
        texts = []
        
        # 新格式（JS传来）
        if 'texts' in result and result['texts']:
            texts = [t for t in result['texts'] if isinstance(t, str)]
        # 旧格式（原始OCR）
        elif result.get('code') == 100 and result.get('data'):
            for item in result['data']:
                if isinstance(item, dict) and 'text' in item:
                    texts.append(item['text'])
        
        if not texts:
            logger.warning("没有可导出的文本内容")
        
        try:
            if format_type.upper() == "TXT":
                # 如果 filename 已包含扩展名，直接使用
                if filename.endswith('.txt'):
                    file_path = os.path.join(output_dir, filename)
                else:
                    file_path = os.path.join(output_dir, f"{filename}.txt")
                
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write('\n'.join(texts))
                
                logger.info(f"TXT 导出成功: {file_path}")
                return file_path
            
            elif format_type.upper() == "JSON":
                if filename.endswith('.json'):
                    file_path = os.path.join(output_dir, filename)
                else:
                    file_path = os.path.join(output_dir, f"{filename}.json")
                
                export_data = {
                    "export_time": datetime.now().isoformat(),
                    "filename": filename,
                    "texts": texts,
                    "count": len(texts),
                    "raw_result": result
                }
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(export_data, f, ensure_ascii=False, indent=2)
                
                logger.info(f"JSON 导出成功: {file_path}")
                return file_path
            
            elif format_type.upper() == "EXCEL":
                try:
                    import openpyxl
                    from openpyxl.styles import Font
                except ImportError:
                    logger.error("需要安装 openpyxl 库来导出 Excel 文件")
                    return None
                
                if filename.endswith('.xlsx'):
                    file_path = os.path.join(output_dir, filename)
                else:
                    file_path = os.path.join(output_dir, f"{filename}.xlsx")
                
                # 使用 write_only 模式：流式写入，性能提升 50-100 倍
                wb = openpyxl.Workbook(write_only=True)
                ws = wb.create_sheet(title="OCR识别结果")
                
                # 设置列宽（必须在写入行之前设置）
                ws.column_dimensions['A'].width = 8
                ws.column_dimensions['B'].width = 60
                
                from openpyxl.cell import WriteOnlyCell
                
                # 写入表头（使用 WriteOnlyCell 实现样式）
                cell1 = WriteOnlyCell(ws, value="序号")
                cell1.font = Font(bold=True)
                cell2 = WriteOnlyCell(ws, value="识别文本")
                cell2.font = Font(bold=True)
                ws.append([cell1, cell2])
                
                # 写入数据（无样式以提升性能）
                for i, text in enumerate(texts, 1):
                    ws.append([i, text])
                
                wb.save(file_path)
                logger.info(f"Excel 导出成功: {file_path}")
                return file_path
            else:
                logger.error(f"不支持的导出格式: {format_type}")
                return None
        except Exception as e:
            logger.error(f"导出失败: {e}", exc_info=True)
            return None
    
    def export_batch(self, results: List[Dict[str, Any]], format_type: str, 
                     file_path: str, column_headers: Optional[List[str]] = None,
                     include_original_text: bool = True) -> Optional[str]:
        """
        批量导出识别结果（支持三种格式，优化输出结构，支持动态列）
        
        Args:
            results: 结果列表，每项是一个字典，键对应列头
            format_type: 导出格式 ("TXT", "JSON", "Excel")
            file_path: 输出文件路径（含扩展名）
            column_headers: 列头列表（可选，None 时使用 results[0] 的键作为列头）
            include_original_text: 是否包含原始文本（Excel 导出时使用）
            
        Returns:
            导出文件路径，失败返回 None
        """
        if not results:
            logger.warning("尝试导出空结果")
            return None
        
        try:
            format_upper = format_type.upper()
            
            if format_upper == "TXT":
                # TXT 格式：清晰区分每个图片的内容，支持动态列
                with open(file_path, 'w', encoding='utf-8') as f:
                    # 动态确定列头
                    if column_headers:
                        headers = column_headers
                    elif results and isinstance(results[0], dict):
                        headers = [k for k in results[0].keys() if not k.startswith('_')]
                    else:
                        headers = ["file_name", "text", "extracted_text"]
                    
                    for i, item in enumerate(results, 1):
                        # 文件头分隔符
                        f.write("=" * 60 + "\n")
                        
                        # 动态写入每个列的内容
                        for header in headers:
                            value = item.get(header, '')
                            # 跳过内部使用的键（如 _column_headers）
                            if header.startswith('_'):
                                continue
                            f.write(f"【{header}】\n")
                            f.write(f"{value}\n\n")
                        
                        f.write("\n")  # 文件间空行
                
                logger.info(f"TXT 批量导出成功: {file_path}")
                return file_path
            
            elif format_upper == "JSON":
                # JSON 格式：使用键值对结构，便于程序处理和人工阅读，支持动态列
                export_data = {
                    "export_time": datetime.now().isoformat(),
                    "total_files": len(results),
                    "results": []
                }
                
                # 动态确定列头
                if column_headers:
                    headers = column_headers
                elif results and isinstance(results[0], dict):
                    headers = [k for k in results[0].keys() if not k.startswith('_')]
                else:
                    headers = ["file_name", "text", "extracted_text"]
                
                for i, item in enumerate(results, 1):
                    entry = {"index": i}
                    
                    # 动态写入每个列的内容
                    for header in headers:
                        # 跳过内部使用的键
                        if header.startswith('_'):
                            continue
                        value = item.get(header, '')
                        entry[header] = value
                    
                    export_data["results"].append(entry)
                
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(export_data, f, ensure_ascii=False, indent=2)
                
                logger.info(f"JSON 批量导出成功: {file_path}")
                return file_path
            
            elif format_upper == "EXCEL":
                # Excel 格式：动态列（与界面表格完全一致）
                # 优化：使用 write_only 模式，性能提升 50-100 倍
                try:
                    import openpyxl
                    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                    from openpyxl.cell import WriteOnlyCell
                except ImportError:
                    logger.error("需要安装 openpyxl 库来导出 Excel 文件")
                    logger.info("请运行: pip install openpyxl")
                    return None
                
                wb = openpyxl.Workbook(write_only=True)
                ws = wb.create_sheet(title="OCR识别结果")
                
                # 动态确定列头
                if column_headers:
                    headers = column_headers
                else:
                    # 使用辅助方法构建列头
                    headers = self._build_excel_headers(results, include_original_text)
                
                # 设置列宽（必须在写入行之前设置）
                for col in range(1, len(headers) + 1):
                    from openpyxl.utils import get_column_letter
                    # 文件名列窄一些，文本列宽一些
                    if headers[col-1] == 'file_name':
                        ws.column_dimensions[get_column_letter(col)].width = 25
                    elif headers[col-1] == 'text':
                        ws.column_dimensions[get_column_letter(col)].width = 50
                    else:
                        ws.column_dimensions[get_column_letter(col)].width = 15
                
                # 设置表头样式
                header_font = Font(bold=True, size=12)
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # 写入表头（使用 WriteOnlyCell 实现样式）
                header_cells = []
                for header in headers:
                    cell = WriteOnlyCell(ws, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
                    header_cells.append(cell)
                ws.append(header_cells)
                
                # 写入数据（无样式以提升性能）
                for item in results:
                    row_data = self._build_excel_row(item, headers, include_original_text)
                    ws.append(row_data)
                
                # 添加统计信息
                stats_cell = WriteOnlyCell(ws, value="统计信息")
                stats_cell.font = Font(bold=True)
                ws.append([stats_cell])
                ws.append([f"总文件数: {len(results)}"])
                ws.append([f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
                
                wb.save(file_path)
                logger.info(f"Excel 批量导出成功: {file_path}")
                return file_path
            
            else:
                logger.error(f"不支持的导出格式: {format_type}")
                return None
        
        except Exception as e:
            logger.error(f"批量导出失败: {e}", exc_info=True)
            return None
    
    def get_combined_text(self, separator: str = '\n') -> str:
        """获取合并后的文本
        
        Args:
            separator: 分隔符
            
        Returns:
            合并后的文本
        """
        texts = []
        for item in self.results:
            result = item['result']
            if result.get('code') == 100 and result.get('data'):
                for line in result['data']:
                    if isinstance(line, dict) and 'text' in line:
                        texts.append(line['text'])
        return separator.join(texts)
    
    def _build_excel_headers(self, results: List[Dict], include_original_text: bool = True) -> List[str]:
        """
        构建 Excel 列头（支持展开 extracted 字段）
        
        Args:
            results: 识别结果列表
            include_original_text: 是否包含原始文本
            
        Returns:
            列头列表
        """
        if not results:
            return ["file_name", "text"]
        
        # 检查是否有 extracted 字段
        has_extracted = any('extracted' in item for item in results)
        
        if has_extracted:
            # 有模板解析结果：展开 extracted 字段
            headers = ["file_name"]
            
            # 获取所有可能出现的 extracted 字段
            extracted_keys = set()
            for item in results:
                extracted = item.get('extracted', {})
                if isinstance(extracted, dict):
                    extracted_keys.update(extracted.keys())
            
            # 添加 extracted 字段作为列
            headers.extend(sorted(extracted_keys))
            
            # 可选：添加原始文本列
            if include_original_text:
                headers.append("text")
            
            return headers
        else:
            # 没有模板解析结果：使用默认列头
            headers = ["file_name"]
            if include_original_text:
                headers.append("text")
            return headers
    
    def _build_excel_row(self, item: Dict, headers: List[str], include_original_text: bool = True) -> List[Any]:
        """
        构建 Excel 数据行（支持展开 extracted 字段）
        
        Args:
            item: 识别结果项
            headers: 列头列表
            include_original_text: 是否包含原始文本
            
        Returns:
            数据行列表
        """
        row_data = []
        
        for header in headers:
            if header == "file_name":
                # 文件名
                row_data.append(item.get("file_name", ""))
            elif header == "text":
                # 原始文本
                if include_original_text:
                    # 从 result 中提取文本
                    result = item.get("result", {})
                    if result.get("code") == 100 and result.get("data"):
                        texts = [line.get("text", "") for line in result["data"] if isinstance(line, dict)]
                        row_data.append("\n".join(texts))
                    else:
                        row_data.append(item.get("text", ""))
                else:
                    row_data.append("")  # 不包含原始文本
            elif header in item.get("extracted", {}):
                # extracted 字段
                row_data.append(item.get("extracted", {}).get(header, ""))
            else:
                # 其他字段
                row_data.append(item.get(header, ""))
        
        return row_data


# 全局导出器实例
_exporter: Optional[ResultExporter] = None


def get_exporter() -> ResultExporter:
    """获取全局导出器实例
    
    Returns:
        导出器实例
    """
    global _exporter
    if _exporter is None:
        logger.info("创建全局导出器实例")
        _exporter = ResultExporter()
    return _exporter


def reset_exporter() -> ResultExporter:
    """重置导出器
    
    Returns:
        新的导出器实例
    """
    global _exporter
    logger.info("重置导出器")
    _exporter = ResultExporter()
    return _exporter    
